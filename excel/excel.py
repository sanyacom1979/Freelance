from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Назначение ширины колонок Excel
def book_col_width(cl, ws):
    col_width = [17, 40, 70, 23, 20, 18, 18, 19, 17]                        # Прикидка ширины колонок
    for i in range(1, cl + 1):
        ws.column_dimensions[get_column_letter(i)].width = col_width[i - 1]


# Прорисовка границ таблички, выравнивание текста, выделение цветом заголовка
def book_styles(rw, cl, ws):
    border_style = Side(border_style="thin", color="000000")
    for i in range(1, rw + 2):                   # учитываем заголовок
        for j in range(1, cl + 1):
            cell = ws[get_column_letter(j) + str(i)]
            cell.border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)
            if i == 1:    # заголовок
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = PatternFill('solid', fgColor="c0c0c0")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)


# Функция записи данных из списка в таблицу Excel
def to_excel(w_list, f_excel):
    wb = Workbook()
    ws = wb.active
    l_len = len(w_list)                         # количество строк
    d_len = len(w_list[0])                      # количество колонок
    for i, l_item in enumerate(w_list):
        if i == 0:
            ws.append(list(l_item.keys()))
        ws.append(list(l_item.values()))
    book_col_width(d_len, ws)
    book_styles(l_len, d_len, ws)
    wb.save(f_excel)
    print(f"Результаты сохранены в файл {f_excel}.")
